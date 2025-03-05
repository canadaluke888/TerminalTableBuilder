from openpyxl import load_workbook, Workbook
import os

class ExcelHandler:

    def __init__(self, table_builder):
        self.table_builder = table_builder


    def load_excel(self, file: str = None) -> None:
        file_name = file or self.table_builder.input_handler.get_user_input("[bold yellow]Enter the path to the Excel file[/]: ")

        if file_name is None:
            return
        
        if not os.path.exists(file_name):
            self.table_builder.system_message.create_error_message("File not found.")
            return

        try:
            wb = load_workbook(filename=file_name, data_only=True)
            sheet = wb.active  # Get the first sheet

            data = [[cell.value for cell in row] for row in sheet.iter_rows()]
            if not data:
                self.table_builder.system_message.create_error_message("The Excel file is empty.")
                return

            self.table_builder.table_data["columns"] = [{"name": col, "type": "str"} for col in data[0]]
            self.table_builder.table_data["rows"] = [{col: row[i] if i < len(row) else "" for i, col in enumerate(data[0])} for row in data[1:]]
            self.table_builder.name = os.path.splitext(os.path.basename(file_name))[0]
            if self.table_builder.settings.get_setting("infer_data_types") == "on":
                self.table_builder.table_specs.infer_column_types()
            self.table_builder.system_message.create_information_message("Excel file loaded successfully.")
        except Exception as e:
            self.table_builder.system_message.create_error_message(f"Failed to load Excel file: {e}")

    def save_excel(self) -> None:
        """
        Save the current table data to an .xlsx file using openpyxl.
        """
        if not self.table_builder.table_data["columns"] or not self.table_builder.table_data["rows"]:
            self.table_builder.system_message.create_error_message("No table data to export to Excel.")
            return
        
        use_table_name = self.table_builder.input_handler.get_user_input("[bold yellow]Use table name as save file name? (y/n)[/]: ").lower().strip()

        if use_table_name == "y":
            file_name = f"{self.table_builder.name}.xlsx"
        elif use_table_name == "n":
            file_name = self.table_builder.input_handler.get_user_input("[bold yellow]Enter the name of the Excel file (without extension)[/]: ").strip() + ".xlsx"
        elif use_table_name is None:
            return
        else:
            self.table_builder.system_message.create_error_message("Invalid input.")
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            # Write column headers
            column_headers = [col["name"] for col in self.table_builder.table_data["columns"]]
            ws.append(column_headers)

            # Write row data
            for row in self.table_builder.table_data["rows"]:
                ws.append([row.get(col, "") for col in column_headers])

            wb.save(file_name)
            self.table_builder.system_message.create_information_message(f"Table data successfully saved to '[bold cyan]{file_name}[/]'.")
        except Exception as e:
            self.table_builder.system_message.create_error_message(f"Failed to save table as Excel file: {e}")