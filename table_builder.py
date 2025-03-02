from rich.table import Table
from message_panel.message_panel import MessagePanel
import csv
import os
import pdfplumber
from autocomplete.autocomplete import Autocomplete
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table as PDFTable, TableStyle
from reportlab.lib import colors
import json
from rich.panel import Panel
from settings.settings import Settings
from rich.console import Console
import sqlite3
from database.database import Database
from openpyxl import load_workbook, Workbook
from pyexcel_ods3 import save_data, get_data

class TableBuilder:
    def __init__(self, console: Console, settings: Settings, database: Database, print_on_start: bool = False, name_on_start: str = None):
        """
        Initialize a new table.

        :param console: Console instance for printing to screen.
        :param settings: Settings instance for using user settings.
        :param database: Database instance for updating the database.
        :param print_on_start: Signal to print the table with the provided CSV file path argument using the CSV flag in CLI mode.
        :param name_on_start: Signal to name the table with the provided argument using the TB flag in CLI mode.
        """
        self.console = console
        self.database = database
        self.autocomplete = Autocomplete(self.console)
        self.settings = settings
        self.message_panel = MessagePanel(self.console)
        if name_on_start:
            self.name = name_on_start
        else:
            self.name = self.name_table()
        self.table_data = {"columns": [], "rows": []}
        self.table_saved = False

        if print_on_start:
            self.print_table()

    def get_user_input(self, prompt: str) -> str:
        """
        Get user input and allow cancellation.
        """
        while True:
            user_input = self.console.input(prompt).strip()
            if user_input == "/cancel":
                return None
            return user_input

    def save_pdf(self):
        """
        Save the current table data to a PDF file.
        """
        if not self.table_data["columns"] or not self.table_data["rows"]:
            self.message_panel.create_error_message("No table data to export to PDF.")
            return

        # Prompt for the file name
        use_table_name = self.console.input(
            "[bold yellow]Use table name as save file name? (y/n)[/]: ").strip().lower()

        if use_table_name == "y":
            file_name = f"{self.name}.pdf"
        elif use_table_name == "n":
            file_name = self.console.input(
                "[bold yellow]Enter the name of the PDF file (without extension)[/]: ").strip() + ".pdf"
        else:
            self.message_panel.create_error_message("Invalid input.")
            return

        try:
            # Create the PDF document
            pdf = SimpleDocTemplate(file_name, pagesize=letter)
            elements = []

            # Prepare the table data for the PDF
            # Extract column names from the dictionaries
            column_headers = [column["name"] for column in self.table_data["columns"]]
            pdf_data = [column_headers]  # Header row

            # Add row data
            for row in self.table_data["rows"]:
                pdf_data.append([row.get(column["name"], "") for column in self.table_data["columns"]])

            # Create the table with styling
            table = PDFTable(pdf_data)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]))

            elements.append(table)

            # Build the PDF
            pdf.build(elements)
            self.table_saved = True
            self.message_panel.create_information_message(
                f"Table data successfully exported to '[bold cyan]{file_name}[/]'."
            )
        except Exception as e:
            self.message_panel.create_error_message(f"Failed to save table as PDF: {e}")

    def load_pdf(self):
        """
        Load table data from a PDF file using pdfplumber.
        """
        file_name = self.console.input("[bold yellow]Enter path to the PDF file[/]: ").strip()

        if not os.path.exists(file_name):
            self.message_panel.create_error_message("File not found.")
            return

        try:
            with pdfplumber.open(file_name) as pdf:
                extracted_data = []
                for page in pdf.pages:
                    tables = page.extract_table()
                    if tables:
                        extracted_data.extend(tables)

            if not extracted_data:
                self.message_panel.create_error_message("No table data found in the PDF.")
                return

            # Use first row as column headers
            self.table_data["columns"] = [{"name": col, "type": "str"} for col in extracted_data[0]]
            self.table_data["rows"] = [{col: row[i] if i < len(row) else "" for i, col in enumerate(extracted_data[0])} for row in extracted_data[1:]]
            self.message_panel.create_information_message("PDF file loaded successfully.")

            if self.settings.get_autoprint_table() == "on":
                self.print_table()
        except Exception as e:
            self.message_panel.create_error_message(f"Failed to load PDF file: {e}")

        
    def ensure_connected_database(self) -> bool:
        """
        Ensure there is a connected database before proceeding with table operations.
        """
        if not self.database.is_connected():
            self.message_panel.create_error_message("No database connected.")
            return False
        return True
        
    def save_to_database(self) -> None:
        """
        Save the current table data to the connected database, with error handling to abort
        on any failure.
        """
        if not self.ensure_connected_database():
            return

        if not self.table_data["columns"]:
            self.message_panel.create_error_message("No columns defined. Add columns before saving.")
            return

        try:
            quoted_table_name = f'"{self.name}"'

            # Check if the table exists
            self.database.cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            existing_tables = [row[0] for row in self.database.cursor.fetchall()]

            if self.name in existing_tables:
                if self.settings.get_auto_update() == "on":
                    # Auto-update: overwrite the existing table without prompting
                    self.database.cursor.execute(f"DROP TABLE {quoted_table_name}")
                else:
                    # Prompt the user for action
                    action = self.console.input(
                        f"[bold yellow]Table '[bold cyan]{self.name}[/]' already exists. "
                        "Do you want to overwrite it or save with a new name? (overwrite/new)[/]: "
                    ).strip().lower()

                    if action == "overwrite":
                        self.database.cursor.execute(f"DROP TABLE {quoted_table_name}")
                    elif action == "new":
                        new_name = self.console.input("[bold yellow]Enter a new name for the table[/]: ").strip()
                        if not new_name:
                            self.message_panel.create_error_message("Table name cannot be empty.")
                            return
                        self.name = new_name
                        quoted_table_name = f'"{self.name}"'
                    else:
                        self.message_panel.create_error_message("Invalid action. Please enter 'overwrite' or 'new'.")
                        return

            # Generate columns definition with data types
            columns_definition = []
            for column in self.table_data["columns"]:
                column_name = f'"{column["name"]}"'
                column_type = column["type"].upper()
                if column_type == "STR":
                    column_type = "TEXT"
                elif column_type == "INT":
                    column_type = "INTEGER"
                elif column_type == "FLOAT":
                    column_type = "REAL"
                elif column_type == "BOOL":
                    column_type = "BOOLEAN"
                else:
                    self.message_panel.create_error_message(
                        f"Unsupported data type for column '[bold cyan]{column['name']}': {column['type']}[/]'"
                    )
                    return

                columns_definition.append(f"{column_name} {column_type}")
            columns_definition_str = ", ".join(columns_definition)

            # Create the table
            self.database.cursor.execute(f"CREATE TABLE {quoted_table_name} ({columns_definition_str})")

            # Insert rows
            for row in self.table_data["rows"]:
                column_names = ", ".join(f'"{col["name"]}"' for col in self.table_data["columns"])
                placeholders = ", ".join("?" for _ in self.table_data["columns"])
                values = [row.get(col["name"], None) for col in self.table_data["columns"]]
                self.database.cursor.execute(
                    f"INSERT INTO {quoted_table_name} ({column_names}) VALUES ({placeholders})", values
                )

            self.database.connection.commit()
            self.table_saved = True
            self.message_panel.create_information_message(
                f"Table '[bold cyan]{self.name}[/]' saved to database '[bold red]{self.database.get_current_database()}[/]'."
            )
        except Exception as e:
            self.message_panel.create_error_message(f"Failed to save table to database: {e}")


            
    def load_from_database(self) -> None:
        """
        Load a table from the connected database, including column data types.
        """
        if not self.ensure_connected_database():
            return

        try:
            self.database.cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = [row[0] for row in self.database.cursor.fetchall()]

            if not tables:
                self.message_panel.create_error_message("No tables found in database.")
                return

            self.console.print("[bold green]Available Tables:[/]")
            for idx, table in enumerate(tables, start=1):
                self.console.print(f"{idx}. {table}")

            table_number = int(self.console.input("[bold yellow]Enter the number of the table to load[/]: ")) - 1
            if not (0 <= table_number < len(tables)):
                self.message_panel.create_error_message("Invalid table number.")
                return

            table_name = tables[table_number]
            quoted_table_name = f'"{table_name}"'

            # Fetch column information
            self.database.cursor.execute(f"PRAGMA table_info({quoted_table_name})")
            columns_info = self.database.cursor.fetchall()

            # Map SQL types back to program types
            sql_to_program_types = {
                "TEXT": "str",
                "INTEGER": "int",
                "REAL": "float",
                "BOOLEAN": "bool"
            }
            columns = [{"name": col[1], "type": sql_to_program_types.get(col[2].upper(), "str")} for col in columns_info]

            # Fetch rows
            self.database.cursor.execute(f"SELECT * FROM {quoted_table_name}")
            raw_rows = self.database.cursor.fetchall()

            # Convert rows to dictionaries and handle boolean conversion
            rows = []
            for raw_row in raw_rows:
                row = {}
                for idx, column in enumerate(columns):
                    value = raw_row[idx]
                    if column["type"] == "bool":
                        value = bool(value)  # Convert 1/0 to True/False
                    row[column["name"]] = value
                rows.append(row)

            self.table_data["columns"] = columns
            self.table_data["rows"] = rows

            self.name = table_name
            self.table_saved = True
            self.message_panel.create_information_message(
                f"Table '[bold cyan]{table_name}[/]' loaded successfully from database '[bold red]{self.database.get_current_database()}[/]'."
            )
            if self.settings.get_autoprint_table() == "on":
                self.print_table()
        except Exception as e:
            self.message_panel.create_error_message(f"Failed to load table: {e}")


    def list_tables(self) -> None:
        """
        List all tables in the connected database.
        """
        if not self.ensure_connected_database():
            return

        try:
            self.database.cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = [row[0] for row in self.database.cursor.fetchall()]

            if tables:
                self.console.print("[bold green]Available Tables:[/]")
                for idx, table in enumerate(tables, start=1):
                    self.console.print(f"{idx}. {table}")
            else:
                self.message_panel.create_error_message("No tables found in the database.")
        except Exception as e:
            self.message_panel.create_error_message(f"Failed to list tables: {e}")

    def get_tables(self) -> list:
        """
        Returns a list of table names in the currently connected database.
        """
        if not self.ensure_connected_database():
            self.message_panel.create_error_message("No database is connected.")
            return []
        
        try:
            self.database.cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            return [row[0] for row in self.database.cursor.fetchall()]
        except sqlite3.Error as e:
            self.message_panel.create_error_message(f"Failed to fetch tables [blue]{e}[/]")

    def next_table_number(self) -> int:
        """
        Returns the next available table number based on the existing tables in the current database.
        """
        existing_tables = self.get_tables()
        return len(existing_tables) + 1

    def show_current_table(self) -> MessagePanel:
        self.message_panel.create_information_message(f"Current Table: [bold cyan]{self.name}[/]")
        
    def save_to_csv(self) -> None:
        """
        Save the current table data to a CSV file.
        """
        use_table_name = self.console.input(
            "[bold yellow]Use table name as save file name? (y/n)[/]: ").lower().strip()

        if use_table_name == "y":
            file_name = f"{self.name}.csv"
        elif use_table_name == "n":
            file_name = self.console.input(
                "[bold yellow]Enter the name of the file (without extension)[/]: ") + ".csv"
        else:
            self.message_panel.create_error_message("Invalid input.")
            return

        # Write table data to the CSV file
        try:
            with open(file_name, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)

                # Write header row (columns)
                if self.table_data["columns"]:
                    writer.writerow(self.table_data["columns"])

                # Write data rows
                for row in self.table_data["rows"]:
                    writer.writerow([row.get(column, "") for column in self.table_data["columns"]])

            self.table_saved = True
            self.message_panel.create_information_message(
                f"Table data successfully saved to '[bold red]{file_name}[/]'."
            )
        except Exception as e:
            self.message_panel.create_error_message(f"Failed to save file: {e}")

    def name_table(self) -> str:
        self.table_saved = False
        return self.console.input("[bold yellow]Enter a name for the new table[/]: ")
    
    def save_to_json(self):
        """Save the table data to a JSON file."""
        use_table_name = self.get_user_input(
            "[bold yellow]Use table name as save file name? (y/n)[/]: ").lower().strip()

        if use_table_name == "y":
            file_name = f"{self.name}.json"
        elif use_table_name == "n":
            file_name = self.console.input(
                "[bold yellow]Enter the name of the file (without extension)[/]: ") + ".json"
            if file_name is None:
                return
        elif file_name is None:
            return
        else:
            self.message_panel.create_error_message("Invalid input.")
            return
        
        with open(file_name, 'w') as f:
            json.dump(self.table_data, f, indent=4)

            self.table_saved = True
            self.message_panel.create_information_message(
                f"Table data successfully saved to '[bold red]{file_name}[/]'."
            )

            
    def load_csv(self, path: str = None) -> None:
        """
        Load a CSV file and update the table data with all columns defaulting to strings.

        Args:
            path (str): Path to the CSV file. If not provided, prompts the user.
        """
        # Prompt for CSV path if not provided through CLI arg.
        csv_path = path or self.get_user_input("[bold yellow]Enter path to CSV file[/]: ").strip()

        if csv_path is None:
            return

        # Validate the file path
        if not os.path.isfile(csv_path):
            self.message_panel.create_error_message("Invalid path or file does not exist.")
            return

        try:
            with open(csv_path, 'r', encoding='utf-8') as csv_file:
                reader = csv.reader(csv_file)
                rows = list(reader)  # Convert CSV reader to a list of rows

                # Ensure the CSV is not empty
                if not rows:
                    self.message_panel.create_error_message("CSV file is empty.")
                    return

                # Use the first row as column names, defaulting to string type
                self.table_data["columns"] = [{"name": col, "type": "str"} for col in rows[0]]

                # Populate the rows with column-value dictionaries
                self.table_data["rows"] = [
                    {col: value for col, value in zip([c["name"] for c in self.table_data["columns"]], row)}
                    for row in rows[1:]
                ]

                # Mark the table as unsaved and notify the user
                self.table_saved = False
                self.message_panel.create_information_message("CSV file loaded successfully.")

                # Automatically print the table if the setting is enabled
                if self.settings.get_autoprint_table() == "on":
                    self.print_table()

        except UnicodeDecodeError:
            self.message_panel.create_error_message("Failed to decode file. Ensure it is UTF-8 encoded.")
        except Exception as e:
            self.message_panel.create_error_message(f"Failed to load CSV file: {e}")


            
    def load_batch_csv(self) -> None:
        """
        Load multiple CSV files from a specified directory into the database.
        """
        directory = self.get_user_input("[bold yellow]Enter the directory of CSV files[/]: ").strip()

        if directory is None:
            return

        # Ensure a database is connected
        if not self.ensure_connected_database():
            self.message_panel.create_error_message("No database is connected")
            return

        # Validate directory path
        if not os.path.exists(directory) or not os.path.isdir(directory):
            self.message_panel.create_error_message("Directory does not exist or is invalid.")
            return

        recursive_load = self.console.input("[bold yellow]Load CSV files from subdirectories? (y/n)[/]: ").strip().lower()
        csv_files = []

        # Collect CSV files
        try:
            if recursive_load == 'y':
                for root, _, files in os.walk(directory):
                    csv_files.extend([os.path.join(root, file) for file in files if file.endswith('.csv')])
            elif recursive_load == 'n':
                csv_files = [os.path.join(directory, file) for file in os.listdir(directory) if file.endswith('.csv')]
            else:
                self.message_panel.create_error_message("Invalid input! Please enter 'y' or 'n'.")
                return
        except Exception as e:
            self.message_panel.create_error_message(f"Error accessing directory: {e}")
            return

        if not csv_files:
            self.message_panel.create_error_message("No CSV files found in the specified directory.")
            return

        self.message_panel.create_information_message(f"Found [bold cyan]{len(csv_files)}[/] CSV files.")


        existing_tables = self.get_tables()  # Get list of existing tables

        for file in csv_files:
            try:
                self.load_csv(path=file)

                base_name = os.path.splitext(os.path.basename(file))[0]  # Use filename as default table name
                if base_name in existing_tables:
                    # Prompt user to rename or use default
                    user_choice = self.console.input(
                        f"[bold yellow]Table '{base_name}' already exists. Enter a new name or press Enter to use default.[/]: "
                    ).strip()

                    if not user_choice:
                        base_name = f"Table {self.database.next_table_number()}"

                self.name = base_name
                self.save_to_database()

                self.message_panel.create_information_message(f"Successfully loaded table '[bold cyan]{self.name}[/]' from file '[bold red]{file}[/]'.")
            except Exception as e:
                self.message_panel.create_error_message(f"Error processing '[bold blue]{file}[/]': {e}")

        self.table_saved = True
        self.message_panel.create_information_message("Batch CSV loading complete!")


    def load_excel(self, file: str = None) -> None:
        file_name = file or self.get_user_input("[bold yellow]Enter the path to the Excel file[/]: ")

        if file_name is None:
            return
        
        if not os.path.exists(file_name):
            self.message_panel.create_error_message("File not found.")
            return

        try:
            wb = load_workbook(filename=file_name, data_only=True)
            sheet = wb.active  # Get the first sheet

            data = [[cell.value for cell in row] for row in sheet.iter_rows()]
            if not data:
                self.message_panel.create_error_message("The Excel file is empty.")
                return

            self.table_data["columns"] = [{"name": col, "type": "str"} for col in data[0]]
            self.table_data["rows"] = [{col: row[i] if i < len(row) else "" for i, col in enumerate(data[0])} for row in data[1:]]
            self.message_panel.create_information_message("Excel file loaded successfully.")

            if self.settings.get_autoprint_table() == "on":
                self.print_table()
        except Exception as e:
            self.message_panel.create_error_message(f"Failed to load Excel file: {e}")


    def save_excel(self) -> None:
        """
        Save the current table data to an .xlsx file using openpyxl.
        """
        if not self.table_data["columns"] or not self.table_data["rows"]:
            self.message_panel.create_error_message("No table data to export to Excel.")
            return
        
        use_table_name = self.get_user_input("[bold yellow]Use table name as save file name? (y/n)[/]: ").lower().strip()

        if use_table_name == "y":
            file_name = f"{self.name}.xlsx"
        elif use_table_name == "n":
            file_name = self.console.input("[bold yellow]Enter the name of the Excel file (without extension)[/]: ").strip() + ".xlsx"
        elif use_table_name is None:
            return
        else:
            self.message_panel.create_error_message("Invalid input.")
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            # Write column headers
            column_headers = [col["name"] for col in self.table_data["columns"]]
            ws.append(column_headers)

            # Write row data
            for row in self.table_data["rows"]:
                ws.append([row.get(col, "") for col in column_headers])

            wb.save(file_name)
            self.message_panel.create_information_message(f"Table data successfully saved to '[bold cyan]{file_name}[/]'.")
        except Exception as e:
            self.message_panel.create_error_message(f"Failed to save table as Excel file: {e}")

    def load_ods(self):
        """
        Load table data from an .ods file using pyexcel-ods3.
        """
        file_name = self.console.input("[bold yellow]Enter path to the ODS file[/]: ").strip()

        if not os.path.exists(file_name):
            self.message_panel.create_error_message("File not found.")
            return

        try:
            data = get_data(file_name)
            sheet_name = list(data.keys())[0]  # Ensure we get the first sheet correctly
            sheet_data = data[sheet_name]

            if not sheet_data or not isinstance(sheet_data, list):
                self.message_panel.create_error_message("The ODS file is empty or has an invalid format.")
                return

            self.table_data["columns"] = [{"name": col, "type": "str"} for col in sheet_data[0]]
            self.table_data["rows"] = [{col: row[i] if i < len(row) else "" for i, col in enumerate(sheet_data[0])} for row in sheet_data[1:]]
            self.message_panel.create_information_message("ODS file loaded successfully.")

            if self.settings.get_autoprint_table() == "on":
                self.print_table()
        except Exception as e:
            self.message_panel.create_error_message(f"Failed to load ODS file: {e}")

    def save_ods(self):
        """
        Save the current table data to an .ods file using pyexcel-ods3.
        """
        if not self.table_data["columns"] or not self.table_data["rows"]:
            self.message_panel.create_error_message("No table data to export to ODS.")
            return

        file_name = self.console.input("[bold yellow]Enter the name of the ODS file (without extension)[/]: ").strip() + ".ods"

        try:
            column_headers = [col["name"] for col in self.table_data["columns"]]
            data = [[col for col in column_headers]]
            data.extend([[row.get(col, "") for col in column_headers] for row in self.table_data["rows"]])

            save_data(file_name, {"Sheet1": data})
            self.message_panel.create_information_message(f"Table data successfully saved to '[bold cyan]{file_name}[/]'.")
        except Exception as e:
            self.message_panel.create_error_message(f"Failed to save table as ODS file: {e}")

    def get_num_columns(self) -> int:
        """
        Returns:
            int: The total amount of columns in the table.
        """
        return len(self.table_data["columns"])

    def get_num_rows(self) -> int:
        """
        Returns:
            int: The total amount of rows in the table.
        """
        return len(self.table_data["rows"])

    def add_column(self) -> None:
        column_name = self.get_user_input("[bold yellow]Enter column name[/]: ")
        if column_name is None:
            return

        if column_name in [col["name"] for col in self.table_data["columns"]]:
            self.message_panel.create_error_message("Column already exists.")
            return

        types = ["int", "float", "str", "bool"]
        self.console.print("[bold green]Available Types:[/]")
        for idx, t in enumerate(types, start=1):
            self.console.print(f"{idx}. {t}")

        while True:
            type_number = self.get_user_input("[bold yellow]Enter the number of the type for this column[/]: ")
            if type_number is None:
                return
            try:
                type_number = int(type_number) - 1
                if not (0 <= type_number < len(types)):
                    self.message_panel.create_error_message("Invalid type number. Try again.")
                    continue
                selected_type = types[type_number]
                break
            except ValueError:
                self.message_panel.create_error_message("Invalid input. Please enter a number.")
                continue

        self.table_data["columns"].append({"name": column_name, "type": selected_type})
        for row in self.table_data["rows"]:
            row[column_name] = ""
        self.table_saved = False
        self.message_panel.create_information_message(f"Column '[bold cyan]{column_name}[/]' added with type '[bold red]{selected_type}'.[/]")

    def change_column_type(self) -> None:
        """
        Change the data type of a specific column in the table.
        """
        if not self.table_data["columns"]:
            self.message_panel.create_error_message("No columns defined. Add columns before changing types.")
            return

        # Display available columns for selection
        self.console.print("[bold green]Available Columns:[/]")
        for idx, column in enumerate(self.table_data["columns"], start=1):
            self.console.print(f"{idx}. {column['name']} (Current Type: {column['type']})")

        try:
            column_number = int(self.console.input("[bold yellow]Enter the number of the column to change type[/]: ")) - 1
            if not (0 <= column_number < len(self.table_data["columns"])):
                self.message_panel.create_error_message("Invalid column number.")
                return
            selected_column = self.table_data["columns"][column_number]
        except ValueError:
            self.message_panel.create_error_message("Invalid input. Please enter a number.")
            return

        # Display available types for selection
        types = ["int", "float", "str", "bool"]
        self.console.print("[bold green]Available Types:[/]")
        for idx, t in enumerate(types, start=1):
            self.console.print(f"{idx}. {t}")

        try:
            type_number = int(self.console.input("[bold yellow]Enter the number of the new type[/]: ")) - 1
            if not (0 <= type_number < len(types)):
                self.message_panel.create_error_message("Invalid type number.")
                return
            new_type = types[type_number]
        except ValueError:
            self.message_panel.create_error_message("Invalid input. Please enter a number.")
            return

        # Confirm the change
        confirm = self.console.input(
            f"[bold red]Are you sure you want to change column '{selected_column['name']}' "
            f"from '{selected_column['type']}' to '{new_type}'? (y/n)[/]: "
        ).strip().lower()

        if confirm == "y":
            # Apply the type change
            selected_column["type"] = new_type
            self.message_panel.create_information_message(
                f"Column '[bold cyan]{selected_column['name']}[/]' type changed to '[bold red]{new_type}'.[/]"
            )
            self.table_saved = False
        else:
            self.message_panel.create_information_message("[bold yellow]Column type change cancelled.[/]")

    def edit_column_name(self) -> None:
        """
        Edits the name of an existing column in the table.
        """
        if not self.table_data["columns"]:
            self.message_panel.create_error_message("No columns defined. Add columns before renaming.")
            return

        # Display available columns for selection
        self.console.print("[bold green]Available Columns:[/]")
        for idx, column in enumerate(self.table_data["columns"], start=1):
            self.console.print(f"{idx}. {column['name']} (Type: {column['type']})")

        try:
            column_number = int(self.console.input("[bold yellow]Enter the number of the column to rename[/]: ")) - 1
            if not (0 <= column_number < len(self.table_data["columns"])):
                self.message_panel.create_error_message("Invalid column number.")
                return
            selected_column = self.table_data["columns"][column_number]
        except ValueError:
            self.message_panel.create_error_message("Invalid input. Please enter a number.")
            return

        # Prompt for the new column name
        new_name = self.console.input(f"[bold yellow]Enter the new name for column '[bold cyan]{selected_column['name']}[/]': [/]").strip()

        # Validate new name (e.g., avoid duplicates)
        if any(column["name"] == new_name for column in self.table_data["columns"]):
            self.message_panel.create_error_message("A column with this name already exists. Please choose a different name.")
            return

        # Apply the name change
        old_name = selected_column["name"]
        selected_column["name"] = new_name

        # Update all row data to reflect the name change
        for row in self.table_data["rows"]:
            row[new_name] = row.pop(old_name, "")

        self.table_saved = False
        self.message_panel.create_information_message(f"Column '[bold cyan]{old_name}[/]' renamed to '[bold green]{new_name}[/]' successfully.")


    def add_row(self) -> None:
        """
        Adds a row to the table data with validation for column data types.
        """
        if not self.table_data["columns"]:
            self.message_panel.create_error_message("No columns defined. Add columns before adding rows.")
            return
        
        row_data = {}
        for column in self.table_data["columns"]:
            column_name = column["name"]
            data_type = column["type"]

            while True:
                cell_data = self.get_user_input(f"[bold yellow]Enter data for column '[bold cyan]{column_name}[/]' ([bold red]{data_type})[/]: ")
                if cell_data is None:
                    return
                if data_type == "int" and cell_data.isdigit():
                    row_data[column_name] = int(cell_data)
                    break
                elif data_type == "float":
                    try:
                        row_data[column_name] = float(cell_data)
                        break
                    except ValueError:
                        self.message_panel.create_error_message("Invalid data. Expected a float.")
                elif data_type == "bool" and cell_data.lower() in ["true", "false"]:
                    row_data[column_name] = cell_data.lower() == "true"
                    break
                elif data_type == "str":
                    row_data[column_name] = cell_data
                    break
                else:
                    self.message_panel.create_error_message("Invalid input. Please enter a valid value.")

        self.table_data["rows"].append(row_data)
        self.table_saved = False
        self.message_panel.create_information_message("Row added with validated data.")


    def edit_cell(self) -> None:
        """
        Edits the cell content using direct indexing.
        """
        if not self.table_data["columns"] or not self.table_data["rows"]:
            self.message_panel.create_error_message("No table data to edit. Add rows and columns first.")
            return

        # Display table cells with their indices
        self.console.print("[bold green]Table Cells:[/]")
        for row_idx, row in enumerate(self.table_data["rows"], start=1):
            row_display = [
                f"({row_idx},{col_idx + 1}) {row[col['name']]}" for col_idx, col in enumerate(self.table_data["columns"])
            ]
            self.console.print(f"Row {row_idx}: " + " | ".join(row_display))

        try:
            # Prompt for cell index
            cell_position = self.get_user_input("[bold yellow]Enter cell position as 'row,column' (e.g., 1,2)[/]: ").strip()
            if cell_position is None:
                return
            row_idx, col_idx = map(int, cell_position.split(","))
            row_idx -= 1  # Convert to 0-based index
            col_idx -= 1  # Convert to 0-based index

            if not (0 <= row_idx < len(self.table_data["rows"]) and 0 <= col_idx < len(self.table_data["columns"])):
                self.message_panel.create_error_message("Invalid cell position.")
                return

            # Fetch column and its type
            column = self.table_data["columns"][col_idx]
            column_name = column["name"]
            column_type = column["type"]

            # Prompt for new value with type validation
            while True:
                new_data = self.console.input(f"[bold yellow]Enter new data for cell ([bold cyan]{row_idx + 1},{col_idx + 1}[/]) ([bold red]{column_name}[/]: [bold blue]{column_type})[/]: ").strip()
                if column_type == "int" and new_data.isdigit():
                    new_value = int(new_data)
                    break
                elif column_type == "float":
                    try:
                        new_value = float(new_data)
                        break
                    except ValueError:
                        self.message_panel.create_error_message("Invalid data. Expected a float.")
                elif column_type == "bool":
                    if new_data.lower() in ["true", "false"]:
                        new_value = new_data.lower() == "true"
                        break
                    else:
                        self.message_panel.create_error_message("Invalid data. Expected 'true' or 'false'.")
                elif column_type == "str":
                    new_value = new_data
                    break
                else:
                    self.message_panel.create_error_message(f"Unsupported data type: [bold cyan]{column_type}[/]")

            # Update the cell
            self.table_data["rows"][row_idx][column_name] = new_value
            self.table_saved = False
            self.message_panel.create_information_message("Cell updated successfully.")

        except ValueError:
            self.message_panel.create_error_message("Invalid input format. Use 'row,column'.")
        except Exception as e:
            self.message_panel.create_error_message(f"Failed to edit cell: {e}")


    def remove_column(self) -> None:
        """
        Removes a column based on the column name provided by the user.
        Aborts if the column name is not found or any error occurs.
        """
        column_name = self.console.input("[bold yellow]Enter column name to remove[/]: ").strip()

        # Check if the column exists
        column_names = [col["name"] for col in self.table_data["columns"]]
        if column_name not in column_names:
            self.message_panel.create_error_message(f"Column '[bold cyan]{column_name}[/]' does not exist.")
            return

        try:
            # Remove the column from the table structure
            self.table_data["columns"] = [
                col for col in self.table_data["columns"] if col["name"] != column_name
            ]
            for row in self.table_data["rows"]:
                row.pop(column_name, None)

            self.table_saved = False
            self.message_panel.create_information_message(f"Column '[bold cyan]{column_name}[/]' removed successfully.")
            
        except Exception as e:
            self.message_panel.create_error_message(f"Failed to remove column: {e}")


    def remove_row(self) -> None:
        """
        Removes a row from the table based on the row index given.
        """
        row_number = int(self.console.input("[bold yellow]Enter row number to remove (1-based index)[/]: ")) - 1
        if 0 <= row_number < len(self.table_data["rows"]):
            self.table_data["rows"].pop(row_number)
            self.table_saved = False
            self.message_panel.create_information_message("Row removed.")
        else:
            self.message_panel.create_error_message("Invalid row number.")
            
    def delete_table(self) -> None:
        """
        Delete a table from the connected database by selecting it from a list of available tables.
        """
        if not self.ensure_connected_database():
            return

        try:
            # Get the list of available tables
            self.database.cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = [row[0] for row in self.database.cursor.fetchall()]

            if not tables:
                self.message_panel.create_error_message("[bold yellow]No tables found in the database.[/]")
                return

            # Display the available tables
            self.console.print("[bold green]Available Tables:[/]")
            for idx, table in enumerate(tables, start=1):
                self.console.print(f"{idx}. {table}")

            # Get user selection
            table_number = int(self.console.input("[bold yellow]Enter the number of the table to delete[/]: ")) - 1

            # Validate selection
            if 0 <= table_number < len(tables):
                table_name = tables[table_number]
                quoted_table_name = f'"{table_name}"'  # Safely quote the table name

                # Confirm deletion
                confirm = self.console.input(f"[bold red]Are you sure you want to delete table '[bold cyan]{table_name}[/]'? (y/n)[/]: ").lower().strip()
                if confirm == 'y':
                    # Execute deletion
                    self.database.cursor.execute(f"DROP TABLE {quoted_table_name}")
                    self.database.connection.commit()
                    self.message_panel.create_information_message(f"Table '[bold cyan]{table_name}[/]' has been deleted successfully.")
                else:
                    self.message_panel.create_information_message("[bold yellow]Table deletion cancelled.[/]")
            else:
                self.message_panel.create_error_message("[bold red]Invalid table number.[/]")
        except Exception as e:
            self.message_panel.create_error_message(f"[bold red]Failed to delete table: {e}[/]")


    def build_table(self) -> Table:
        """
        Takes the current table data and builds the table with it.

        Returns:
            Table: The rendered table with all of the data.
        """
        if not self.table_data["columns"]:
            self.message_panel.create_error_message("No columns defined. Add columns before building the table.")
            return Table(border_style="yellow", show_lines=True)

        # Create a Rich Table instance
        table = Table(title=self.name, border_style="yellow", show_lines=True)

        # Add columns with type information
        for column in self.table_data["columns"]:
            column_name = column["name"]
            column_type = column["type"]
            table.add_column(f"{column_name} ([bold red]{column_type}[/])", style="cyan")

        # Add rows
        for row in self.table_data["rows"]:
            row_values = [str(row.get(column["name"], "")) for column in self.table_data["columns"]]
            table.add_row(*row_values, style="magenta")

        return table


    def print_table(self) -> None:
        """
        Prints the built table to the screen.
        """
        table = self.build_table()
        self.console.print(table)

    def print_table_data(self) -> Panel:
        """
        Prints the table data to the screen.
        """
        self.console.print(Panel(str(self.table_data), title="[bold red]Table Data[/]", title_align="center", border_style="cyan"))

    def clear_table(self) -> None:
        """
        Clears the table data.
        """
        self.table_data = {"columns": [], "rows": []}
        self.message_panel.create_information_message("Table cleared.")

    def launch_builder(self) -> None:
        """
        Launches the table builder loop.
        """

        if self.settings.get_hide_instructions() == "off":
            self.message_panel.print_table_builder_instructions()

        while True:
            builder_command = self.console.input("[bold red]Table Builder[/] - [bold yellow]Enter a command[/]: ").lower().strip()

            if builder_command == "print help":
                self.message_panel.print_table_builder_instructions()

            elif builder_command == "add column":
                self.add_column()

                if self.settings.get_autoprint_table() == "on":
                    self.print_table()

                if self.settings.get_auto_update() == "on":
                    self.save_to_database()

            elif builder_command == "change type":
                self.change_column_type()
                if self.settings.get_autoprint_table() == "on":
                    self.print_table()

                if self.settings.get_auto_update() == "on":
                    self.save_to_database()

            elif builder_command == "rename column":
                self.edit_column_name()

                if self.settings.get_autoprint_table() == "on":
                    self.print_table()

                if self.settings.get_auto_update() == "on":
                    self.save_to_database()

            elif builder_command == "add row":
                self.add_row()
                if self.settings.get_autoprint_table() == "on":
                    self.print_table()

                if self.settings.get_auto_update() == "on":
                    self.save_to_database()

            elif builder_command == "edit cell":
                self.edit_cell()
                if self.settings.get_autoprint_table() == "on":
                    self.print_table()

                if self.settings.get_auto_update() == "on":
                    self.save_to_database()

            elif builder_command == "remove column":
                self.remove_column()
                if self.settings.get_autoprint_table() == "on":
                    self.print_table()

                if self.settings.get_auto_update() == "on":
                    self.save_to_database()

            elif builder_command == "remove row":
                self.remove_row()
                if self.settings.get_autoprint_table() == "on":
                    self.print_table()

                if self.settings.get_auto_update() == "on":
                    self.save_to_database()

            elif builder_command == "print table":
                self.print_table()

            elif builder_command == "print table data":
                self.print_table_data()

            elif builder_command == "current table":
                self.show_current_table()

            elif builder_command == "clear table":
                self.clear_table()

            elif builder_command == "rename":
                self.name = self.name_table()
                if self.settings.get_autoprint_table() == "on":
                    self.print_table()

            elif builder_command == "load table":
                self.load_from_database()
                
            elif builder_command == "save table":
                self.save_to_database()

            elif builder_command == "delete table":
                self.delete_table()
                
            elif builder_command == "load csv":
                self.load_csv()

            elif builder_command == "load xl":
                self.load_excel()

            elif builder_command == "load ods":
                self.load_ods()
                
            elif builder_command == "load csv batch":
                self.load_batch_csv()
            
            elif builder_command == "list tables":
                self.list_tables()

            elif builder_command == "save csv":
                self.save_to_csv()

            elif builder_command == "save xl":
                self.save_excel()

            elif builder_command == "save ods":
                self.save_ods()
                
            elif builder_command == "save pdf":
                self.save_pdf()
            
            elif builder_command == "load pdf":
                self.load_pdf()

            elif builder_command == "save json":
                self.save_to_json()

            elif builder_command == "help":
                self.message_panel.print_table_builder_instructions()

            elif builder_command == "exit":
                if not self.table_saved:
                    exit_response = self.console.input("[bold red]Are you sure you want to exit without saving? (y/n)[/]: ").lower().strip()
                    
                    if exit_response == "y":
                        break
                    elif exit_response == "n":
                        continue
                    else:
                        self.message_panel.create_error_message("Invalid input.")
                        
                else:
                    break

            else:
                self.message_panel.create_error_message("Invalid Command.")
                self.autocomplete.suggest_command(builder_command, self.autocomplete.table_builder_commands)
